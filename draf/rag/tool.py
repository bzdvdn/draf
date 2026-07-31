"""RAG tool — retrieve context from a vector store for LLM use."""

import csv
import glob
import os
from typing import Callable

from draf.tool.tool import Tool
from draf.rag.base import VectorStore
from draf.rag.embedder import Embedder
from draf.rag.chunker import Chunker


def load_documents_csv(
    path: str, text_column: str = "text", delimiter: str = ","
) -> list[tuple[str, dict]]:
    """Load documents from a CSV file.

    Returns a list of ``(text, metadata)`` tuples. The *text_column*
    column becomes the document text; all other columns become metadata.
    """
    docs: list[tuple[str, dict]] = []
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter=delimiter)
        for row in reader:
            text = row.get(text_column, "")
            if not text:
                continue
            meta = {k: v for k, v in row.items() if k != text_column and v != ""}
            docs.append((text, meta))
    return docs


def load_documents_txt(path: str, encoding: str = "utf-8") -> list[tuple[str, dict]]:
    """Load documents from ``.txt`` files (glob supported).

    Each matched file becomes one ``(text, {"id": stem, "path": ...})`` doc.
    """
    paths = sorted(glob.glob(path)) if glob.has_magic(path) else [path]
    docs: list[tuple[str, dict]] = []
    for p in paths:
        with open(p, encoding=encoding) as f:
            text = f.read()
        if not text.strip():
            continue
        stem = os.path.splitext(os.path.basename(p))[0]
        docs.append((text, {"id": stem, "path": p}))
    return docs


def load_documents_pdf(path: str) -> list[tuple[str, dict]]:
    """Load text from a PDF, one document per page (requires ``pypdf``)."""
    try:
        from pypdf import PdfReader
    except ImportError as e:
        msg = "install 'draf[rag-pdf]' (pypdf) to load PDF documents"
        raise ImportError(msg) from e

    reader = PdfReader(path)
    stem = os.path.splitext(os.path.basename(path))[0]
    docs: list[tuple[str, dict]] = []
    for i, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        if text.strip():
            docs.append((text, {"id": f"{stem}_{i}", "path": path, "page": i}))
    return docs


def load_documents_excel(
    path: str,
    text_column: str = "text",
    sheet: int | str = 0,
    skip_header: bool = True,
) -> list[tuple[str, dict]]:
    """Load documents from an Excel file, one per row (requires ``openpyxl``)."""
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        msg = "install 'draf[rag-excel]' (openpyxl) to load Excel documents"
        raise ImportError(msg) from e

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if isinstance(sheet, str) else wb.worksheets[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    if skip_header:
        header = [str(c) if c is not None else "" for c in rows[0]]
        data_rows = rows[1:]
    else:
        header = [str(i) for i in range(len(rows[0]))]
        data_rows = rows

    stem = os.path.splitext(os.path.basename(path))[0]
    docs: list[tuple[str, dict]] = []
    for i, row in enumerate(data_rows, start=1):
        record = {header[j]: ("" if v is None else str(v)) for j, v in enumerate(row)}
        text = record.get(text_column, "")
        if not text:
            continue
        record.setdefault("id", f"{stem}_{i}")
        meta = {k: v for k, v in record.items() if k != text_column}
        docs.append((text, meta))
    return docs


_DOCUMENT_LOADERS: dict[str, Callable[..., list[tuple[str, dict]]]] = {
    "csv": load_documents_csv,
    "txt": load_documents_txt,
    "pdf": load_documents_pdf,
    "excel": load_documents_excel,
}


class RAGTool(Tool):
    """Tool that searches a vector store and returns ranked results.

    Usage::

        store = InMemoryVectorStore(dim=768)
        embedder = Embedder(provider="openai")
        tool = RAGTool(store, embedder)
        await tool.add_document("some long text")
        result = await tool.arun(query="find this")

    Can also be built from a config dict (e.g. a ``tools:`` entry in a
    workflow YAML)::

        {
          "embedder": {"provider": "ollama", "model": "nomic-embed-text"},
          "store": {"type": "in_memory", "dim": 768},
          "documents": [
            {"type": "csv", "path": "docs.csv"},
            {"type": "txt", "path": "corpus/*.txt"},
            {"type": "pdf", "path": "manual.pdf"},
            {"type": "excel", "path": "table.xlsx", "text_column": "content"},
          ],
        }

    Supported document types (loaders): ``csv``, ``txt`` (glob), ``pdf``
    (``draf[rag-pdf]``), ``excel`` (``draf[rag-excel]``). Supported store
    types: ``in_memory`` (default), ``sqlite`` (stdlib file persistence),
    ``chroma`` / ``qdrant`` / ``pgvector`` (via ``draf[embedding]``).
    ``documents`` may also be a bare path (CSV shorthand) or a list of
    inline ``{"id": ..., "text": ...}`` dicts. Documents are embedded
    lazily on the first search.
    """

    name = "rag"
    description = "Search documents using RAG"

    def __init__(
        self,
        config: dict | None = None,
        *,
        store: VectorStore | None = None,
        embedder: Embedder | None = None,
        chunker: Chunker | None = None,
        documents: list[tuple[str, dict]] | None = None,
    ):
        self.store = store
        self.embedder = embedder
        self.chunker = chunker or Chunker()
        self._documents: list[tuple[str, dict]] = list(documents or [])
        self._seeded = False
        if isinstance(config, dict):
            self._apply_config(config)

    def _apply_config(self, config: dict) -> None:
        emb = config.get("embedder") or {}
        self.embedder = Embedder(
            provider=emb.get("provider", "ollama"),
            model=emb.get("model", "nomic-embed-text"),
            base_url=emb.get("base_url"),
        )

        store_cfg = config.get("store") or {}
        store_type = store_cfg.get("type", "in_memory")
        if store_type == "in_memory":
            from draf.rag.stores import InMemoryVectorStore

            self.store = InMemoryVectorStore(dim=store_cfg.get("dim", 768))
        elif store_type == "sqlite":
            from draf.rag.stores import SQLiteVectorStore

            self.store = SQLiteVectorStore(
                path=store_cfg.get("path", "./vectors.db"),
                dim=store_cfg.get("dim"),
            )
        elif store_type == "chroma":
            from draf.rag.stores import ChromaVectorStore

            self.store = ChromaVectorStore(
                path=store_cfg.get("path", "./chroma"),
                collection=store_cfg.get("collection", "draf"),
            )
        elif store_type == "qdrant":
            from draf.rag.stores import QdrantVectorStore

            self.store = QdrantVectorStore(
                host=store_cfg.get("host", "localhost"),
                port=store_cfg.get("port", 6333),
                collection=store_cfg.get("collection", "draf"),
            )
        elif store_type == "pgvector":
            from draf.rag.stores import PGVectorStore

            self.store = PGVectorStore(
                dsn=store_cfg.get("dsn", ""),
                table=store_cfg.get("table", "draf_vectors"),
            )
        else:
            msg = f"unsupported store type: {store_type}"
            raise ValueError(msg)

        if config.get("chunker"):
            self.chunker = Chunker(**config["chunker"])

        documents = config.get("documents", [])
        if isinstance(documents, str):
            self._load_source({"path": documents})
        elif isinstance(documents, dict):
            self._load_source(documents)
        else:
            for doc in documents or []:
                if "text" in doc:
                    meta = {k: v for k, v in doc.items() if k != "text"}
                    self._documents.append((doc["text"], meta))
                else:
                    self._load_source(doc)

    def _load_source(self, cfg: dict) -> None:
        stype = cfg.get("type", "csv")
        loader = _DOCUMENT_LOADERS.get(stype)
        if loader is None:
            msg = f"unsupported document type: {stype}"
            raise ValueError(msg)
        kwargs = {k: v for k, v in cfg.items() if k != "type"}
        if "file" in kwargs and "path" not in kwargs:
            kwargs["path"] = kwargs.pop("file")
        self._documents.extend(loader(**kwargs))

    async def _ensure_seeded(self) -> None:
        if not self._seeded and self._documents:
            await self.add_documents(self._documents)
        self._seeded = True

    async def arun(self, query: str = "", k: int = 5) -> str:  # type: ignore[override]
        """Search documents and return formatted results."""
        await self._ensure_seeded()
        assert self.embedder is not None
        assert self.store is not None
        query_vec = await self.embedder.embed(query)
        results = await self.store.search(query_vec, k=k)
        if not results:
            return ""
        context_parts = []
        for i, (doc_id, score, meta) in enumerate(results):
            text = meta.get("text", doc_id)
            context_parts.append(f"[{i + 1}] (score: {score:.3f}) {text}")
        return "\n\n".join(context_parts)

    async def add_document(self, text: str, metadata: dict | None = None) -> None:
        """Chunk, embed, and store a document."""
        metadata = metadata or {}
        assert self.embedder is not None
        assert self.store is not None
        chunks = self.chunker.chunk(text)
        vectors = []
        embeddings = await self.embedder.embed_many(chunks)
        for i, (chunk, vec) in enumerate(zip(chunks, embeddings)):
            doc_id = (
                f"{metadata.get('id', 'doc')}_{i}"
                if metadata.get("id")
                else f"chunk_{i}"
            )
            meta = {**metadata, "text": chunk, "chunk_index": i}
            vectors.append((doc_id, vec, meta))
        await self.store.add(vectors)

    async def add_documents(self, docs: list[tuple[str, dict]]) -> None:
        """Add multiple documents at once."""
        for text, meta in docs:
            await self.add_document(text, meta)
