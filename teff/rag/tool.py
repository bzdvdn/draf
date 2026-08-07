"""RAG tool — retrieve context from a vector store for LLM use."""

import csv
import glob
import os
import uuid
from typing import Callable

from teff.rag.base import VectorStore
from teff.rag.chunker import Chunker
from teff.rag.embedder import Embedder, embedder_from_config
from teff.tool.tool import Tool


def load_documents_csv(
    path: str, text_column: str = "text", delimiter: str = ","
) -> list[tuple[str, dict]]:
    """Load documents from a CSV file.

    Returns a list of ``(text, metadata)`` tuples. The *text_column*
    column becomes the document text; all other columns become metadata.
    """
    docs: list[tuple[str, dict]] = []
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter=delimiter)
        for row in reader:
            text = row.get(text_column, "")
            if not text:
                continue
            meta = {k: v for k, v in row.items() if k != text_column and v != ""}
            docs.append((text, meta))
    return docs


def load_documents_txt(path: str, encoding: str = "utf-8") -> list[tuple[str, dict]]:
    """Load documents from ``.txt`` files (glob supported).

    Each matched file becomes one ``(text, {"id": stem, "path": ...})`` doc.
    """
    paths = sorted(glob.glob(path)) if glob.has_magic(path) else [path]
    docs: list[tuple[str, dict]] = []
    for p in paths:
        with open(p, encoding=encoding) as f:
            text = f.read()
        if not text.strip():
            continue
        stem = os.path.splitext(os.path.basename(p))[0]
        docs.append((text, {"id": stem, "path": p}))
    return docs


def load_documents_pdf(path: str) -> list[tuple[str, dict]]:
    """Load text from a PDF, one document per page (requires ``pypdf``)."""
    try:
        from pypdf import PdfReader
    except ImportError as e:
        msg = "install 'teff[rag-pdf]' (pypdf) to load PDF documents"
        raise ImportError(msg) from e

    reader = PdfReader(path)
    stem = os.path.splitext(os.path.basename(path))[0]
    docs: list[tuple[str, dict]] = []
    for i, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        if text.strip():
            docs.append((text, {"id": f"{stem}_{i}", "path": path, "page": i}))
    return docs


def load_documents_excel(
    path: str,
    text_column: str = "text",
    sheet: int | str = 0,
    skip_header: bool = True,
) -> list[tuple[str, dict]]:
    """Load documents from an Excel file, one per row (requires ``openpyxl``)."""
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        msg = "install 'teff[rag-excel]' (openpyxl) to load Excel documents"
        raise ImportError(msg) from e

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if isinstance(sheet, str) else wb.worksheets[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    if skip_header:
        header = [str(c) if c is not None else "" for c in rows[0]]
        data_rows = rows[1:]
    else:
        header = [str(i) for i in range(len(rows[0]))]
        data_rows = rows

    stem = os.path.splitext(os.path.basename(path))[0]
    docs: list[tuple[str, dict]] = []
    for i, row in enumerate(data_rows, start=1):
        record = {header[j]: ("" if v is None else str(v)) for j, v in enumerate(row)}
        text = record.get(text_column, "")
        if not text:
            continue
        record.setdefault("id", f"{stem}_{i}")
        meta = {k: v for k, v in record.items() if k != text_column}
        docs.append((text, meta))
    return docs


_DOCUMENT_LOADERS: dict[str, Callable[..., list[tuple[str, dict]]]] = {
    "csv": load_documents_csv,
    "txt": load_documents_txt,
    "pdf": load_documents_pdf,
    "excel": load_documents_excel,
}


def _estimate_tokens(text: str) -> int:
    """Rough token estimate (~4 characters per token)."""
    return max(1, (len(text) + 3) // 4)


def _truncate_to_tokens(text: str, tokens: int) -> str:
    """Truncate *text* to an approximate token budget."""
    if _estimate_tokens(text) <= tokens:
        return text
    return text[: tokens * 4]


class RAGTool(Tool):
    """Tool that searches a vector store and returns ranked results.

    Usage::

        store = InMemoryVectorStore(dim=768)
        embedder = Embedder(provider="openai")
        tool = RAGTool(store, embedder)
        await tool.add_document("some long text")
        result = await tool.arun(query="find this")

    Can also be built from a config dict (e.g. a ``tools:`` entry in a
    workflow YAML)::

        {
          "name": "rag_docs",  # optional; overrides the default "rag"
          "embedder": {"provider": "ollama", "model": "nomic-embed-text"},
          "store": {"type": "in_memory", "dim": 768},
          "documents": [
            {"type": "csv", "path": "docs.csv"},
            {"type": "txt", "path": "corpus/*.txt"},
            {"type": "pdf", "path": "manual.pdf"},
            {"type": "excel", "path": "table.xlsx", "text_column": "content"},
          ],
          "filter": {"topic": "news"},        # metadata filter (DSL below)
          "similarity_threshold": 0.5,         # drop low-score hits
          "max_tokens": 1024,                  # context token budget
          "hybrid": true,                      # keyword + semantic blend
          "parent_chunks": true,               # keep full parent text per chunk
          "parent_retrieval": true,            # return whole parent documents
        }

    Supported document types (loaders): ``csv``, ``txt`` (glob), ``pdf``
    (``teff[rag-pdf]``), ``excel`` (``teff[rag-excel]``). Supported store
    types: ``in_memory`` (default), ``sqlite`` (stdlib file persistence),
    ``faiss``, ``lance``, ``chroma``, ``qdrant``, ``milvus``, ``weaviate``,
    ``pgvector``, ``pinecone`` (via ``teff[embedding]``).
    ``documents`` may also be a bare path (CSV shorthand) or a list of
    inline ``{"id": ..., "text": ...}`` dicts. Documents are embedded
    lazily on the first search.

    Filter DSL: ``{"category": "news"}`` (equality), ``{"category":
    ["news", "tech"]}`` (membership), plus ``"$and"`` / ``"$or"`` keys
    combining sub-filters.

    ``parent_chunks`` stores each chunk with a ``parent_id`` and the full
    ``parent_text``; with ``parent_retrieval`` enabled, search returns
    whole parent documents (deduplicated) instead of individual chunks —
    the "small-to-big" pattern.
    """

    name = "rag"
    description = "Search documents using RAG"

    def __init__(
        self,
        config: dict | None = None,
        *,
        store: VectorStore | None = None,
        embedder: Embedder | None = None,
        chunker: Chunker | None = None,
        documents: list[tuple[str, dict]] | None = None,
        name: str | None = None,
        filter: dict | None = None,
        similarity_threshold: float | None = None,
        max_tokens: int | None = None,
        hybrid: bool = False,
        parent_chunks: bool = False,
        parent_retrieval: bool = False,
    ):
        self.store = store
        self.embedder = embedder
        self.chunker = chunker or Chunker()
        self._documents: list[tuple[str, dict]] = list(documents or [])
        self._seeded = False
        self._filters: dict | None = None
        self._threshold: float | None = None
        self._max_tokens: int | None = None
        self._hybrid = False
        self._parent_chunks = False
        self._parent_retrieval = False
        if isinstance(config, dict):
            self._apply_config(config)
        if filter is not None:
            self._filters = filter
        if similarity_threshold is not None:
            self._threshold = similarity_threshold
        if max_tokens is not None:
            self._max_tokens = max_tokens
        if hybrid:
            self._hybrid = True
        if parent_chunks:
            self._parent_chunks = True
        if parent_retrieval:
            self._parent_retrieval = True
        if name is not None:
            self.name = name

    def _apply_config(self, config: dict) -> None:
        if config.get("name"):
            self.name = config["name"]
        self.embedder = embedder_from_config(config)

        store_cfg = config.get("store") or {}
        store_type = store_cfg.get("type", "in_memory")
        if store_type == "in_memory":
            from teff.rag.stores import InMemoryVectorStore

            self.store = InMemoryVectorStore(dim=store_cfg.get("dim", 768))
        elif store_type == "sqlite":
            from teff.rag.stores import SQLiteVectorStore

            self.store = SQLiteVectorStore(
                path=store_cfg.get("path", "./vectors.db"),
                dim=store_cfg.get("dim"),
            )
        elif store_type == "chroma":
            from teff.rag.stores import ChromaVectorStore

            self.store = ChromaVectorStore(
                path=store_cfg.get("path", "./chroma"),
                collection=store_cfg.get("collection", "teff"),
            )
        elif store_type == "qdrant":
            from teff.rag.stores import QdrantVectorStore

            self.store = QdrantVectorStore(
                host=store_cfg.get("host", "localhost"),
                port=store_cfg.get("port", 6333),
                collection=store_cfg.get("collection", "teff"),
            )
        elif store_type == "pgvector":
            from teff.rag.stores import PGVectorStore

            self.store = PGVectorStore(
                dsn=store_cfg.get("dsn", ""),
                table=store_cfg.get("table", "teff_vectors"),
            )
        elif store_type == "faiss":
            from teff.rag.stores import FAISSVectorStore

            self.store = FAISSVectorStore(
                dim=store_cfg.get("dim", 1536),
                path=store_cfg.get("path"),
            )
        elif store_type in ("lance", "lancedb"):
            from teff.rag.stores import LanceVectorStore

            self.store = LanceVectorStore(
                path=store_cfg.get("path", "./lance"),
                table=store_cfg.get("table", "vectors"),
                dim=store_cfg.get("dim"),
            )
        elif store_type == "milvus":
            from teff.rag.stores import MilvusVectorStore

            self.store = MilvusVectorStore(
                uri=store_cfg.get("uri", "./milvus.db"),
                token=store_cfg.get("token", ""),
                collection=store_cfg.get("collection", "teff"),
                dim=store_cfg.get("dim"),
            )
        elif store_type == "weaviate":
            from teff.rag.stores import WeaviateVectorStore

            self.store = WeaviateVectorStore(
                collection=store_cfg.get("collection", "teff"),
                embedded=bool(store_cfg.get("embedded", False)),
                host=store_cfg.get("host", "localhost"),
                http_port=store_cfg.get("http_port", 8080),
                http_secure=bool(store_cfg.get("http_secure", False)),
                grpc_port=store_cfg.get("grpc_port", 50051),
                grpc_secure=bool(store_cfg.get("grpc_secure", False)),
                api_key=store_cfg.get("api_key", ""),
                headers=store_cfg.get("headers"),
                dim=store_cfg.get("dim"),
            )
        elif store_type == "pinecone":
            from teff.rag.stores import PineconeVectorStore

            self.store = PineconeVectorStore(
                index_name=store_cfg.get("index_name", "teff"),
                api_key=store_cfg.get("api_key", ""),
                host=store_cfg.get("host", ""),
                namespace=store_cfg.get("namespace", ""),
                dim=store_cfg.get("dim"),
            )
        else:
            msg = f"unsupported store type: {store_type}"
            raise ValueError(msg)

        if config.get("chunker"):
            self.chunker = Chunker(**config["chunker"])

        self._filters = config.get("filter") or config.get("filters")
        self._threshold = config.get("similarity_threshold")
        self._max_tokens = config.get("max_tokens")
        self._hybrid = bool(config.get("hybrid", False))
        self._parent_chunks = bool(config.get("parent_chunks", False))
        self._parent_retrieval = bool(config.get("parent_retrieval", False))

        documents = config.get("documents", [])
        if isinstance(documents, str):
            self._load_source({"path": documents})
        elif isinstance(documents, dict):
            self._load_source(documents)
        else:
            for doc in documents or []:
                if "text" in doc:
                    meta = {k: v for k, v in doc.items() if k != "text"}
                    self._documents.append((doc["text"], meta))
                else:
                    self._load_source(doc)

    def _load_source(self, cfg: dict) -> None:
        stype = cfg.get("type", "csv")
        loader = _DOCUMENT_LOADERS.get(stype)
        if loader is None:
            msg = f"unsupported document type: {stype}"
            raise ValueError(msg)
        kwargs = {k: v for k, v in cfg.items() if k != "type"}
        if "file" in kwargs and "path" not in kwargs:
            kwargs["path"] = kwargs.pop("file")
        self._documents.extend(loader(**kwargs))

    async def _ensure_seeded(self) -> None:
        if not self._seeded and self._documents:
            await self.add_documents(self._documents)
        self._seeded = True

    async def arun(  # type: ignore[override]
        self,
        query: str = "",
        k: int = 5,
        filter: dict | None = None,
        similarity_threshold: float | None = None,
        max_tokens: int | None = None,
        parent_retrieval: bool | None = None,
    ) -> str:
        """Search documents and return formatted results.

        Any optional argument overrides the value from the config for
        this call; ``None`` falls back to the configured default.
        """
        await self._ensure_seeded()
        assert self.embedder is not None
        assert self.store is not None

        eff_filter = filter if filter is not None else self._filters
        eff_threshold = (
            similarity_threshold
            if similarity_threshold is not None
            else self._threshold
        )
        eff_max_tokens = max_tokens if max_tokens is not None else self._max_tokens
        eff_parent = (
            parent_retrieval if parent_retrieval is not None else self._parent_retrieval
        )

        k_raw = max(k, k * 4) if eff_parent else k
        query_vec = await self.embedder.embed(query)
        results = await self.store.search(
            query_vec,
            k=k_raw,
            filter=eff_filter,
            hybrid=self._hybrid,
            query_text=query,
        )

        if eff_threshold is not None:
            results = [r for r in results if r[1] >= eff_threshold]

        if eff_parent:
            parents: dict[str, tuple[str, float]] = {}
            for doc_id, score, meta in results:
                pid = meta.get("parent_id")
                if pid is None:
                    continue
                parent_text = meta.get("parent_text") or meta.get("text", doc_id)
                if pid not in parents or score > parents[pid][1]:
                    parents[pid] = (parent_text, score)
            ranked = sorted(parents.items(), key=lambda kv: kv[1][1], reverse=True)
            results = [
                (pid, score, {"text": text}) for pid, (text, score) in ranked[:k]
            ]

        context_parts: list[str] = []
        total_tokens = 0
        for doc_id, score, meta in results:
            text = meta.get("text", doc_id)
            if eff_max_tokens is not None:
                tokens = _estimate_tokens(text)
                if total_tokens >= eff_max_tokens:
                    break
                if total_tokens + tokens > eff_max_tokens:
                    remaining = eff_max_tokens - total_tokens
                    text = _truncate_to_tokens(text, remaining)
                    total_tokens = eff_max_tokens
                else:
                    total_tokens += tokens
            context_parts.append(
                f"[{len(context_parts) + 1}] (score: {score:.3f}) {text}"
            )
        return "\n\n".join(context_parts)

    async def add_document(self, text: str, metadata: dict | None = None) -> None:
        """Chunk, embed, and store a document."""
        metadata = metadata or {}
        assert self.embedder is not None
        assert self.store is not None
        chunks = self.chunker.chunk(text)
        parent_id = metadata.get("id")
        if self._parent_chunks:
            parent_id = parent_id or f"doc_{uuid.uuid4().hex[:8]}"
            base_meta = {**metadata, "id": parent_id}
        else:
            base_meta = metadata
        vectors = []
        embeddings = await self.embedder.embed_many(chunks)
        for i, (chunk, vec) in enumerate(zip(chunks, embeddings)):
            if self._parent_chunks:
                doc_id = f"{parent_id}_{i}"
                meta = {
                    **base_meta,
                    "parent_id": parent_id,
                    "parent_text": text,
                    "text": chunk,
                    "chunk_index": i,
                }
            elif metadata.get("id"):
                doc_id = f"{metadata['id']}_{i}"
                meta = {**metadata, "text": chunk, "chunk_index": i}
            else:
                doc_id = f"chunk_{i}"
                meta = {**metadata, "text": chunk, "chunk_index": i}
            vectors.append((doc_id, vec, meta))
        await self.store.add(vectors)

    async def add_documents(self, docs: list[tuple[str, dict]]) -> None:
        """Add multiple documents at once."""
        for text, meta in docs:
            await self.add_document(text, meta)
